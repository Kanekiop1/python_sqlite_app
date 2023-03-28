# This is a sample Python script.
import sqlite3
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = Workbook()
worksheet = wb.active
worksheet.title = "Expenses"
con = sqlite3.connect('Expenses.sqlite')
cur = con.cursor()
now = datetime.now()
date_time = now.strftime("%Y-%m-%d")
dest_filename = ('%s.xlsx' % date_time)



mysel = cur.execute(
    "SELECT Amount, Category.Name, Date FROM Expense INNER JOIN Category ON Expense.CategoryId = Category.CategoryID ")
for i, row in enumerate(mysel):
    for j, values in enumerate(row):
        worksheet.cell(row=i + 1, column=j + 1).value = values

wb.save(dest_filename)
wb.close()
